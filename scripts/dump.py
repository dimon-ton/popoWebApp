import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
import warnings
warnings.filterwarnings('ignore')

wb = openpyxl.load_workbook('ภาษาอังกฤษ   ป.1.xlsx', data_only=False)
wb_v = openpyxl.load_workbook('ภาษาอังกฤษ   ป.1.xlsx', data_only=True)

target = sys.argv[1] if len(sys.argv) > 1 else None
max_rows = int(sys.argv[2]) if len(sys.argv) > 2 else 25
max_cols = int(sys.argv[3]) if len(sys.argv) > 3 else 15

names = wb.sheetnames
sheets_to_dump = [target] if target else names

for name in sheets_to_dump:
    if name not in names:
        # try by index
        try:
            idx = int(name)
            name = names[idx]
        except:
            print(f'Sheet not found: {name}')
            continue
    ws = wb[name]
    ws_v = wb_v[name]
    print(f'\n===== SHEET: [{name}] dim={ws.dimensions} =====')
    rows = min(ws.max_row, max_rows)
    cols = min(ws.max_column, max_cols)
    # Header column letters
    hdr = '     |' + '|'.join(f'{openpyxl.utils.get_column_letter(c):>14}' for c in range(1, cols+1))
    print(hdr)
    for r in range(1, rows+1):
        row_cells = []
        for c in range(1, cols+1):
            cell = ws.cell(row=r, column=c)
            cell_v = ws_v.cell(row=r, column=c)
            v = cell.value
            if v is None:
                disp = ''
            elif isinstance(v, str) and v.startswith('='):
                vv = cell_v.value
                disp = f'{v[:10]}|{vv}'
            else:
                disp = str(v)
            # truncate
            if len(disp) > 14:
                disp = disp[:13] + '…'
            row_cells.append(f'{disp:>14}')
        print(f'{r:>4} |' + '|'.join(row_cells))

    # Merged cells
    if ws.merged_cells.ranges:
        print(f'\n  MERGED: {len(ws.merged_cells.ranges)} ranges')
        for mr in list(ws.merged_cells.ranges)[:10]:
            print(f'    {mr}')

    # Data validations
    if ws.data_validations.dataValidation:
        print(f'\n  DATA VALIDATIONS: {len(ws.data_validations.dataValidation)}')
        for dv in ws.data_validations.dataValidation[:5]:
            print(f'    type={dv.type} formula={dv.formula1} ranges={dv.sqref}')

    # Defined names referencing this sheet (workbook-level only after loop)
