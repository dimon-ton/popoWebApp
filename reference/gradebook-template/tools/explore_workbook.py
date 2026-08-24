"""List worksheets and dimensions in an Excel grade-book template."""

from argparse import ArgumentParser
from pathlib import Path

import openpyxl


DEFAULT_WORKBOOK = Path(__file__).resolve().parent.parent / "ภาษาอังกฤษ-ป1.xlsx"


def main() -> None:
    parser = ArgumentParser(description=__doc__)
    parser.add_argument("workbook", nargs="?", type=Path, default=DEFAULT_WORKBOOK)
    args = parser.parse_args()

    workbook = openpyxl.load_workbook(args.workbook, data_only=False)
    print("=== SHEET NAMES ===")
    for index, name in enumerate(workbook.sheetnames, start=1):
        worksheet = workbook[name]
        print(
            f"{index}. [{name}] | dim: {worksheet.dimensions} | "
            f"rows: {worksheet.max_row} cols: {worksheet.max_column}"
        )


if __name__ == "__main__":
    main()
