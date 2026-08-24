"""Print a compact view of one or all worksheets in a grade-book workbook."""

from argparse import ArgumentParser
from pathlib import Path
import warnings

import openpyxl


DEFAULT_WORKBOOK = Path(__file__).resolve().parent.parent / "ภาษาอังกฤษ-ป1.xlsx"


def resolve_sheets(workbook, selector):
    if selector is None:
        return workbook.sheetnames
    if selector in workbook.sheetnames:
        return [selector]
    try:
        return [workbook.sheetnames[int(selector)]]
    except (ValueError, IndexError) as error:
        raise SystemExit(f"Sheet not found: {selector}") from error


def main() -> None:
    parser = ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--sheet", help="Worksheet name or zero-based index; defaults to all")
    parser.add_argument("--max-rows", type=int, default=25)
    parser.add_argument("--max-columns", type=int, default=15)
    args = parser.parse_args()

    warnings.filterwarnings("ignore")
    workbook = openpyxl.load_workbook(args.workbook, data_only=False)
    values_workbook = openpyxl.load_workbook(args.workbook, data_only=True)

    for name in resolve_sheets(workbook, args.sheet):
        worksheet = workbook[name]
        values_worksheet = values_workbook[name]
        print(f"\n===== SHEET: [{name}] dim={worksheet.dimensions} =====")
        row_count = min(worksheet.max_row, args.max_rows)
        column_count = min(worksheet.max_column, args.max_columns)
        header = "     |" + "|".join(
            f"{openpyxl.utils.get_column_letter(column):>14}"
            for column in range(1, column_count + 1)
        )
        print(header)

        for row in range(1, row_count + 1):
            cells = []
            for column in range(1, column_count + 1):
                value = worksheet.cell(row=row, column=column).value
                if value is None:
                    display = ""
                elif isinstance(value, str) and value.startswith("="):
                    computed = values_worksheet.cell(row=row, column=column).value
                    display = f"{value[:10]}|{computed}"
                else:
                    display = str(value)
                if len(display) > 14:
                    display = display[:13] + "…"
                cells.append(f"{display:>14}")
            print(f"{row:>4} |" + "|".join(cells))

        merged_ranges = list(worksheet.merged_cells.ranges)
        if merged_ranges:
            print(f"\n  MERGED: {len(merged_ranges)} ranges")
            for merged_range in merged_ranges[:10]:
                print(f"    {merged_range}")

        validations = worksheet.data_validations.dataValidation
        if validations:
            print(f"\n  DATA VALIDATIONS: {len(validations)}")
            for validation in validations[:5]:
                print(
                    f"    type={validation.type} formula={validation.formula1} "
                    f"ranges={validation.sqref}"
                )


if __name__ == "__main__":
    main()
