"""List distinct formulas and workbook-defined names for one worksheet."""

from argparse import ArgumentParser
from pathlib import Path

import openpyxl


DEFAULT_WORKBOOK = Path(__file__).resolve().parent.parent / "ภาษาอังกฤษ-ป1.xlsx"


def main() -> None:
    parser = ArgumentParser(description=__doc__)
    parser.add_argument("sheet", help="Worksheet name")
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    args = parser.parse_args()

    workbook = openpyxl.load_workbook(args.workbook, data_only=False)
    if args.sheet not in workbook.sheetnames:
        raise SystemExit(f"Sheet not found: {args.sheet}")

    worksheet = workbook[args.sheet]
    print(f"=== Formulas in [{args.sheet}] ===")
    seen_formulas = set()
    for row in worksheet.iter_rows():
        for cell in row:
            value = cell.value
            if isinstance(value, str) and value.startswith("="):
                key = value[:60]
                if key not in seen_formulas:
                    seen_formulas.add(key)
                    print(f"  {cell.coordinate}: {value}")

    print("\n=== All defined names ===")
    for name in workbook.defined_names:
        print(f"  {name}: {workbook.defined_names[name].value}")


if __name__ == "__main__":
    main()
