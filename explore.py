import sys
import io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

import openpyxl
wb = openpyxl.load_workbook('ภาษาอังกฤษ   ป.1.xlsx', data_only=False)

print('=== SHEET NAMES ===')
for i, name in enumerate(wb.sheetnames):
    ws = wb[name]
    print(f'{i+1}. [{name}] | dim: {ws.dimensions} | rows: {ws.max_row} cols: {ws.max_column}')
